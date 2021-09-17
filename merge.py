from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import csv

def addtocsv(arr):
    with open('clash.csv', 'a+') as f:
      
    # using csv.writer method from CSV package
        write = csv.writer(f)
        for i in arr:
         write.writerow([i])
    
    f.close()

if __name__ == '__main__':
    arr =[]
    wb = load_workbook('final_record.xlsx')
    ws = wb.active
    for row in range(1,422199):
        match = ws['B' + str(row)].value
        new = match.split('}')
        x = len(new) - 1
        print(row, x)
        arr.append(x)

    addtocsv(arr)
#ws.save('claim_details.xlsx')
