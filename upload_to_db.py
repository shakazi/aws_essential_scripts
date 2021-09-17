from pprint import pprint
import boto3
import openpyxl
import time
import csv

def put_object(fileHash, request='', today = int(time.time()), dynamodb=None):
    if not dynamodb:
        dynamodb = boto3.resource('dynamodb')
    
    table = dynamodb.Table('image-reuse-image-hash-dev')
    response = table.put_item(
       Item={        
                'fileHash': fileHash,
                'createdOn': today,
                'requests': request,
                'updatedOn': today
        }
    )
    return response

def addtocsv(data):
    file = open('final_test.csv', 'a+', newline ='')
    # with file:    
    #     write = csv.writer(file)
    #     write.writerows(data)
    writer = csv.writer(file)
    for key, value in data.items():
        writer.writerow([key, value])

    file.close()

dict1 = {}

def append_to_dict(fileHash, request):
    if fileHash in dict1:
        a = dict1[fileHash]
        a = a + request
        dict1[fileHash]= a
    else:
        dict1[fileHash]= request

if __name__ == '__main__':
    today = int(time.time())
    wb= openpyxl.load_workbook('final_db_data.xlsx')
    print('Workbook loaded!')
    sh1 = wb['Sheet1']
    for i in range (2,640901):
            fileHash = sh1.cell(i,1).value
            request= [
                {
                "sourceElementId": sh1.cell(i,2).value,
                "clientId": "BACKFILL",
                "subLob": sh1.cell(i,4).value,
                "sourceSystem": "ICAN",
                "createdOn": today,
                "lob": "motor"
                }
                ]
            append_to_dict(fileHash,request)
            #output = put_object(fileHash, request, today)
            print("Put object succeeded for item",i, fileHash)
            #pprint(output, sort_dicts=False)
    #print(dict1)
    addtocsv(dict1)
